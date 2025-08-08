# -*- coding: utf-8 -*-
"""
File: workbook.py
Author: Your Name / Tên của bạn
Description: Chứa class Workbook để đại diện và quản lý một file Excel.

--- CHANGELOG ---
Version 0.6.0 (2025-08-08):
    - Tối ưu hóa hiệu suất cho chế độ xóa an toàn (safe=True) bằng cách kết hợp
      openpyxl để tìm công thức và xlwings để thay thế giá trị, tránh treo máy với file lớn.

Version 0.5.0 (2025-08-08):
    - Nâng cấp lớn cho các phương thức quản lý Named Range và External Link.

Version 0.4.0 (2025-08-08):
    - Thêm phương thức .delete_all_named_ranges() và .break_external_links() cơ bản.

Version 0.3.0 (2025-08-08):
    - Nâng cấp .delete_sheet() và .delete_hidden_sheets() với tùy chọn 'safe=True'.

Version 0.2.0 (2025-08-08):
    - Thêm các phương thức nâng cao: .for_each_sheet(), quản lý Named Range, bảo vệ, tính toán.

Version 0.1.0 (2025-08-08):
    - Khởi tạo class Workbook với các chức năng cơ bản.
-------------------
"""

from pathlib import Path
import time
import re
from .sheet import Sheet
from .range import Range
# Thêm thư viện openpyxl để tối ưu hóa việc đọc file
try:
    from openpyxl import load_workbook
except ImportError:
    print("WARNING: 'openpyxl' is not installed. The safe delete feature will be slower. Please install it using: pip install openpyxl")
    load_workbook = None


class Workbook:
    """
    Đại diện cho một file Excel (sổ làm việc).
    """
    def __init__(self, xlw_book, app_instance):
        self._xlw_book = xlw_book
        self._app = app_instance

    def __repr__(self):
        return f"<Workbook [{self.name}]>"

    # --- Properties ---
    @property
    def name(self):
        return self._xlw_book.name

    @property
    def path(self):
        return Path(self._xlw_book.fullname)

    @property
    def app(self):
        return self._app

    @property
    def sheets(self):
        return [Sheet(s, self) for s in self._xlw_book.sheets]

    @property
    def visible_sheets(self):
        """Trả về một danh sách chỉ các sheet đang được hiển thị."""
        return [Sheet(s, self) for s in self._xlw_book.sheets if s.api.Visible == -1]

    @property
    def hidden_sheets(self):
        """Trả về một danh sách chỉ các sheet đang bị ẩn."""
        return [Sheet(s, self) for s in self._xlw_book.sheets if s.api.Visible != -1]

    @property
    def sheet_names(self):
        """Trả về một list tên của tất cả các sheet."""
        return [s.name for s in self._xlw_book.sheets]

    # --- File Lifecycle & Calculation ---
    def save(self):
        self._xlw_book.save()
        return self
        
    def save_as(self, new_path):
        """Lưu workbook với một tên mới."""
        print(f"INFO: Đang lưu workbook thành '{new_path}'...")
        self._xlw_book.save(new_path)
        return self

    def close(self, save_changes=False):
        """Đóng workbook."""
        print(f"INFO: Đang đóng workbook '{self.name}'...")
        if save_changes:
            self.save()
        self._xlw_book.close()

    def activate(self):
        """Kích hoạt (đưa lên phía trước) workbook này."""
        print(f"INFO: Đang kích hoạt workbook '{self.name}'...")
        self._xlw_book.activate()
        return self

    def calculate(self):
        """Buộc Excel tính toán lại tất cả các công thức trong workbook."""
        print(f"INFO: Đang tính toán lại công thức cho workbook '{self.name}'...")
        self._xlw_book.api.Calculate()
        return self

    # --- Protection ---
    def protect(self, password=None):
        """Bảo vệ cấu trúc của workbook (ngăn thêm, xóa, di chuyển sheet)."""
        print(f"INFO: Đang bảo vệ workbook '{self.name}'...")
        self._xlw_book.api.Protect(Password=password)
        return self

    def unprotect(self, password=None):
        """Mở khóa bảo vệ cấu trúc của workbook."""
        print(f"INFO: Đang mở khóa bảo vệ cho workbook '{self.name}'...")
        self._xlw_book.api.Unprotect(Password=password)
        return self

    # --- Sheet Management ---
    def sheet(self, specifier):
        try:
            return Sheet(self._xlw_book.sheets[specifier], self)
        except Exception:
            return None
    
    def add_sheet(self, name, before=None, after=None):
        new_xlw_sheet = self._xlw_book.sheets.add(name, before=self._xlw_book.sheets[before] if before else None, after=self._xlw_book.sheets[after] if after else None)
        return Sheet(new_xlw_sheet, self)

    def delete_sheet(self, specifier, safe=False):
        """
        Xóa một sheet theo tên hoặc index.

        Args:
            specifier (str or int): Tên hoặc index của sheet cần xóa.
            safe (bool, optional): Nếu True, sẽ tìm và phá vỡ các công thức
                                   tham chiếu đến sheet này trước khi xóa.
                                   Mặc định là False.
        """
        try:
            sheet_to_delete = self._xlw_book.sheets[specifier]
            sheet_name_to_delete = sheet_to_delete.name

            if safe:
                self._break_links_to_sheet_optimized(sheet_name_to_delete)

            self.app._app.display_alerts = False
            sheet_to_delete.delete()
            self.app._app.display_alerts = True
            print(f"SUCCESS: Đã xóa thành công sheet '{sheet_name_to_delete}'.")
        except Exception as e:
            print(f"ERROR: Không thể xóa sheet '{specifier}'. Lỗi: {e}")
            self.app._app.display_alerts = True
        return self

    def delete_hidden_sheets(self, safe=False):
        """Xóa tất cả các sheet đang bị ẩn."""
        hidden_names = [sheet.name for sheet in self.hidden_sheets]
        if not hidden_names:
            print("INFO: Không có sheet ẩn nào để xóa.")
            return self

        print(f"INFO: Chuẩn bị xóa {len(hidden_names)} sheet ẩn. Chế độ an toàn: {safe}.")

        if safe:
            self._break_links_to_sheet_optimized(hidden_names)
        
        for name in hidden_names:
            self.delete_sheet(name, safe=False)
        
        print(f"SUCCESS: Đã hoàn tất việc xóa các sheet ẩn.")
        return self
        
    def _break_links_to_sheet_optimized(self, sheets_to_delete_names):
        """
        (Hàm nội bộ tối ưu) Sử dụng openpyxl để nhanh chóng tìm các công thức
        tham chiếu đến các sheet sắp bị xóa, sau đó dùng xlwings để thay thế chúng.
        """
        if not load_workbook:
            print("WARNING: Không thể thực hiện xóa an toàn tối ưu vì thiếu 'openpyxl'. Chuyển sang phương thức chậm hơn.")
            # Nếu không có openpyxl, dùng lại cách cũ
            if isinstance(sheets_to_delete_names, str):
                sheets_to_delete_names = [sheets_to_delete_names]
            for name in sheets_to_delete_names:
                 self._break_links_to_sheet_slow(name)
            return

        print("INFO (Safe Mode Optimized): Đang tìm và phá vỡ các liên kết...")
        
        if isinstance(sheets_to_delete_names, str):
            sheets_to_delete_names = [sheets_to_delete_names]

        try:
            # Lưu file tạm thời để openpyxl đọc được trạng thái mới nhất
            self.save()
            
            # Dùng openpyxl để quét file
            wb_op = load_workbook(self.path, data_only=False)
            cells_to_fix = []
            
            sheets_to_keep_names = [s.title for s in wb_op.worksheets if s.title not in sheets_to_delete_names]

            for sheet_name in sheets_to_keep_names:
                ws = wb_op[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f': # Chỉ kiểm tra ô có công thức
                            formula_str = str(cell.value)
                            for name_to_delete in sheets_to_delete_names:
                                # Kiểm tra xem tên sheet sắp xóa có trong công thức không
                                if f"'{name_to_delete}'!" in formula_str or f"{name_to_delete}!" in formula_str:
                                    cells_to_fix.append({'sheet': sheet_name, 'coord': cell.coordinate})
                                    break # Đã tìm thấy, không cần kiểm tra các tên khác cho ô này

            if not cells_to_fix:
                print("INFO: Không tìm thấy công thức nào cần phá vỡ liên kết.")
                return

            # Dùng xlwings để thay thế giá trị
            print(f"INFO: Tìm thấy {len(cells_to_fix)} công thức cần phá vỡ. Bắt đầu thay thế...")
            for item in cells_to_fix:
                try:
                    cell_to_fix = self.sheet(item['sheet']).range(item['coord'])
                    cell_to_fix.value = cell_to_fix.value
                    print(f"  - Đã thay thế công thức tại {item['sheet']}!{item['coord']}")
                except Exception as e:
                    print(f"  - Lỗi khi thay thế tại {item['sheet']}!{item['coord']}: {e}")

        except Exception as e:
            print(f"ERROR: Đã xảy ra lỗi trong quá trình xóa an toàn tối ưu. Lỗi: {e}")

    def _break_links_to_sheet_slow(self, sheet_name_to_delete):
        """(Hàm dự phòng) Dùng xlwings thuần túy để phá vỡ liên kết. Chậm hơn."""
        print(f"INFO (Safe Mode Slow): Đang tìm và phá vỡ các liên kết đến sheet '{sheet_name_to_delete}'...")
        sheets_to_keep = [s for s in self.sheets if s.name != sheet_name_to_delete]
        
        for sheet in sheets_to_keep:
            for cell in sheet.used_range._xlw_range:
                if cell.has_formula:
                    if f"'{sheet_name_to_delete}'!" in cell.formula or f"{sheet_name_to_delete}!" in cell.formula:
                        cell.value = cell.value

    def for_each_sheet(self, action, include=None, exclude=None):
        """Thực thi một hành động trên nhiều sheet."""
        target_sheets = self.sheets
        
        if include:
            target_sheets = [s for s in target_sheets if s.name in include]
        
        if exclude:
            target_sheets = [s for s in target_sheets if s.name not in exclude]

        print(f"INFO: Đang thực thi hành động trên {len(target_sheets)} sheet...")
        for sheet in target_sheets:
            try:
                action(sheet)
            except Exception as e:
                print(f"ERROR: Lỗi khi thực thi hành động trên sheet '{sheet.name}'. Lỗi: {e}")
        
        return self

    # --- Named Range & Link Management ---
    def _is_valid_named_range(self, name_str):
        """(Hàm nội bộ) Kiểm tra xem một tên có hợp lệ để xử lý hay không."""
        if name_str.startswith('_xlfn'):
            return False
        if not bool(re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', name_str)):
            return False
        if re.match(r'^[A-Za-z]{3}\d', name_str):
            return False
        return True

    def get_named_ranges(self):
        """Lấy danh sách chi tiết tất cả các Named Range trong workbook."""
        names_list = []
        for name in self._xlw_book.api.Names:
            name_info = {
                'name': name.Name,
                'refers_to': name.RefersTo,
                'scope': name.Parent.Name,
                'is_visible': name.Visible
            }
            names_list.append(name_info)
        return names_list

    def delete_all_named_ranges(self, broken_only=False, keep_print_areas=True):
        """
        Xóa nhiều Named Range cùng lúc với các bộ lọc thông minh.

        Args:
            broken_only (bool): Nếu True, chỉ xóa các Named Range bị lỗi #REF!.
            keep_print_areas (bool): Nếu True, sẽ không xóa các Named Range của hệ thống
                                     như 'Print_Area' và 'Print_Titles'.
        """
        if broken_only:
            print("INFO: Đang xóa các Named Range bị lỗi (#REF!)...")
        else:
            print("WARNING: Đang xóa các Named Range hợp lệ...")

        names_to_delete = []
        for name in self._xlw_book.api.Names:
            name_str = name.Name.split('!')[-1] # Lấy tên không bao gồm scope
            refers_to_str = str(name.RefersTo)
            
            # Điều kiện để thêm vào danh sách xóa
            should_delete = False
            if self._is_valid_named_range(name_str):
                if broken_only:
                    if '#REF!' in refers_to_str:
                        should_delete = True
                else: # Xóa tất cả (trừ vùng in nếu được yêu cầu)
                    if keep_print_areas and ("Print_Area" in name_str or "Print_Titles" in name_str):
                        pass # Bỏ qua, không xóa
                    else:
                        should_delete = True
            
            if should_delete:
                names_to_delete.append(name.Name)

        if not names_to_delete:
            print("INFO: Không tìm thấy Named Range nào để xóa.")
            return self

        print(f"INFO: Tìm thấy {len(names_to_delete)} Named Range để xóa. Bắt đầu xóa...")
        for name_str in names_to_delete:
            try:
                self._xlw_book.api.Names(name_str).Delete()
                print(f"  - Đã xóa: {name_str}")
            except Exception as e:
                print(f"  - Lỗi khi xóa {name_str}: {e}")
        
        return self

    def get_external_links(self):
        """Lấy danh sách các nguồn liên kết ngoài."""
        try:
            links = self._xlw_book.api.LinkSources(1) # 1 = xlExcelLinks
            return list(links) if links else []
        except Exception:
            return []

    def break_external_links(self):
        """
        Tìm và phá vỡ tất cả các liên kết đến các file Excel khác một cách an toàn.
        """
        print("INFO: Đang tìm và phá vỡ các liên kết ngoài...")
        links = self.get_external_links()
        
        if not links:
            print("INFO: Không tìm thấy liên kết ngoài nào.")
            return self

        successful_breaks = []
        unsuccessful_breaks = []
        
        for link in links:
            try:
                self._xlw_book.api.BreakLink(Name=link, Type=1)
                print(f"  - Đã phá vỡ liên kết đến: {link}")
                successful_breaks.append(link)
            except Exception as e:
                print(f"  - Lỗi khi phá vỡ liên kết {link}: {e}")
                unsuccessful_breaks.append(link)
        
        print(f"SUCCESS: Hoàn tất. Thành công: {len(successful_breaks)}, Thất bại: {len(unsuccessful_breaks)}.")
        return self

    # --- Conversion & Publishing ---
    def to_pdf(self, output_path=None, quality='standard'):
        if not output_path:
            output_path = self.path.with_suffix('.pdf')
        else:
            output_path = Path(output_path)
        
        self.activate()
        time.sleep(1)
        quality_val = 0 if quality == 'standard' else 1
        self._xlw_book.api.ExportAsFixedFormat(0, str(output_path), Quality=quality_val)
        return self
